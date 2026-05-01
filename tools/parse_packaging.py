#!/usr/bin/env python3
"""
Parse de "Maten + gewichten.xlsx" naar packaging-database.json.

Excel-structuur per sheet (categorie):
  Kolom A: productnaam (of "doos 1" / "Doos 2" voor multi-box producten)
  Kolom B: afmeting "75*72*65" of "200x100x20" (LxBxH in cm)
  Kolom C: gewicht "61 KG"

Multi-box producten: een rij met alleen een naam (geen dims/gewicht), gevolgd
door één of meer "doos N"-rijen met dims+gewicht.

Run:
    python3 tools/parse_packaging.py --in "/path/to/Maten + gewichten.xlsx"

Schrijft `packaging-database.json` in de repo-root.
"""
import argparse, json, os, re, sys
import openpyxl

RE_DIMS   = re.compile(r"(\d+(?:[.,]\d+)?)\s*[x×*]\s*(\d+(?:[.,]\d+)?)\s*[x×*]\s*(\d+(?:[.,]\d+)?)", re.I)
RE_WEIGHT = re.compile(r"(\d+(?:[.,]\d+)?)\s*kg", re.I)
RE_SUBBOX = re.compile(r"^(doos|box|deel|part)\s*\d", re.I)


def parse_dims(s):
    if not s: return None
    m = RE_DIMS.search(str(s))
    if not m: return None
    n = lambda x: float(x.replace(",", "."))
    return {"l": n(m.group(1)), "w": n(m.group(2)), "h": n(m.group(3))}


def parse_weight(s):
    if not s: return None
    m = RE_WEIGHT.search(str(s))
    return float(m.group(1).replace(",", ".")) if m else None


def is_subbox_label(name):
    n = (name or "").strip().rstrip(":")
    return bool(RE_SUBBOX.match(n))


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    categories = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().startswith("blad"):
            continue
        ws = wb[sheet_name]
        products = []
        current = None
        for r in range(2, ws.max_row + 1):
            name_raw = (str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else "")
            dims     = parse_dims(ws.cell(r, 2).value)
            weight   = parse_weight(ws.cell(r, 3).value)

            # Lege rij
            if not name_raw and not dims and not weight:
                if current and current["boxes"]:
                    products.append(current); current = None
                continue

            # Single-box product (naam + dims + weight, niet "doos N")
            if name_raw and dims and weight and not is_subbox_label(name_raw):
                if current and current["boxes"]: products.append(current)
                products.append({"name": name_raw.rstrip(":"),
                                 "boxes": [{"dims": dims, "weight": weight}]})
                current = None
                continue

            # Multi-box parent (alleen naam, geen dims/weight)
            if name_raw and not dims and not is_subbox_label(name_raw):
                if current and current["boxes"]: products.append(current)
                current = {"name": name_raw.rstrip(":"), "boxes": []}
                continue

            # Sub-box (heeft dims of weight)
            if dims or weight:
                if not current:
                    current = {"name": name_raw.rstrip(":") or f"(zonder naam {sheet_name}:R{r})",
                               "boxes": []}
                box = {"label": name_raw.rstrip(":") if name_raw else f"doos {len(current['boxes'])+1}"}
                if dims: box["dims"] = dims
                if weight: box["weight"] = weight
                current["boxes"].append(box)

        if current and current["boxes"]:
            products.append(current)
        # Drop products zonder dozen
        products = [p for p in products if p["boxes"]]
        categories.append({"category": sheet_name.strip(), "products": products})
    return {"categories": categories}


def main():
    ap = argparse.ArgumentParser()
    default_in = os.path.expanduser("~/Downloads/Maten + gewichten .xlsx")
    ap.add_argument("--in", dest="src", default=default_in, help="Pad naar de Excel")
    args = ap.parse_args()

    if not os.path.isfile(args.src):
        print(f"FOUT: niet gevonden: {args.src}", file=sys.stderr); sys.exit(1)

    data = parse_workbook(args.src)
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    out = os.path.join(repo_root, "packaging-database.json")
    with open(out, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    total_p = sum(len(c["products"]) for c in data["categories"])
    total_b = sum(sum(len(p["boxes"]) for p in c["products"]) for c in data["categories"])
    print(f"  ✓ {os.path.relpath(out, repo_root)}")
    print(f"  {total_p} producten, {total_b} dozen, {len(data['categories'])} categorieën")
    for c in data["categories"]:
        n_box = sum(len(p["boxes"]) for p in c["products"])
        print(f"    {c['category']:18}: {len(c['products'])} producten ({n_box} dozen)")


if __name__ == "__main__":
    main()
