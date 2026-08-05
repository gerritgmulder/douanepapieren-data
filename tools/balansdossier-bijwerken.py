#!/usr/bin/env python3
"""Voegt de nieuwe analyses als tabblad toe aan het balansdossier van de accountant.

Kevin krijgt liever een dossier dat aangroeit dan vijf losse CSV's naast elkaar.
Dit script leest de CSV's die de node-tools opleveren en zet ze als extra
tabblad in `Fonteyn - balansdossier 31-12-2025.xlsx`. Bestaat een tabblad al,
dan wordt het vervangen, zodat opnieuw draaien geen dubbele tabbladen geeft.

Gebruik: python3 tools/balansdossier-bijwerken.py
"""
import csv
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAP = Path("/Users/gmulder/Desktop/Intenza/Klanten : Prospects/Fonteyn/"
           "Dashboard/Finance/Geld-goederenbeweging")
DOSSIER = MAP / "Fonteyn - balansdossier 31-12-2025.xlsx"

# tabbladnaam -> (csv-bestand, toelichting bovenaan het blad)
BLADEN = [
    ("1630 per boekjaar", "1630 openstaande posten per boekjaar.csv",
     "Openstaande inkoopleveringen op grootboek 1630 per 31-12-2025, per boekjaar van de eerste boeking. "
     "'in Logic4' = NEE betekent dat het leveringnummer niet meer bestaat; die posten kunnen nooit aflopen."),
    ("Goederenbeweging 2025", "Geld-goederenbeweging 2025.csv",
     "Verloop per voorraadrekening over 2025: beginstand, bijboekingen, afboekingen en eindstand. "
     "Rechtstreeks uit het grootboek, niet uit een export."),
    ("Afboekingen naar soort", "Voorraadafboekingen 2025 naar soort.csv",
     "De 40,3 miljoen aan afboekingen op de voorraad, ingedeeld naar wat de boekingsomschrijving zegt."),
    ("Afboekingen tegenrekening", "Voorraadafboekingen 2025 naar tegenrekening.csv",
     "Dezelfde afboekingen naar tegenrekening. LET OP: alleen de kolom 'exact toegewezen' is hard. "
     "Logic4 boekt per dag een verzamelboeking, en binnen zo'n boeking is per regel niet vast te stellen "
     "welke debetregel bij welke creditregel hoort. De kolom 'naar rato verdeeld' is dus een benadering "
     "en geen bewijs van een tegenboeking."),
]

GROEN = "144734"
KOPVUL = PatternFill("solid", fgColor="E8EFEA")


def getal(waarde):
    """'1.234,56' -> 1234.56 ; laat andere tekst met rust."""
    t = str(waarde).strip()
    if not t:
        return waarde
    kaal = t.replace(".", "").replace(",", ".")
    try:
        if any(c.isdigit() for c in kaal):
            return float(kaal) if ("." in kaal or "," in t) else int(kaal)
    except ValueError:
        pass
    return waarde


def zet_blad(wb, naam, csv_pad, toelichting):
    if naam in wb.sheetnames:
        del wb[naam]
    ws = wb.create_sheet(naam)

    ws["A1"] = toelichting
    ws["A1"].font = Font(italic=True, size=9, color="555555")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 46

    with open(csv_pad, encoding="utf-8-sig") as f:
        rijen = list(csv.reader(f, delimiter=";"))
    if not rijen:
        return 0

    for i, kop in enumerate(rijen[0], start=1):
        c = ws.cell(row=3, column=i, value=kop)
        c.font = Font(bold=True, color=GROEN)
        c.fill = KOPVUL
    for r, rij in enumerate(rijen[1:], start=4):
        for i, waarde in enumerate(rij, start=1):
            ws.cell(row=r, column=i, value=getal(waarde))

    ws.freeze_panes = "A4"
    for i in range(1, len(rijen[0]) + 1):
        breedte = max(len(str(rij[i - 1])) for rij in rijen[:400] if len(rij) >= i)
        ws.column_dimensions[get_column_letter(i)].width = min(max(breedte + 2, 11), 60)
    return len(rijen) - 1


def main():
    if not DOSSIER.exists():
        raise SystemExit(f"dossier niet gevonden: {DOSSIER}")
    # Nooit zonder terugvalbestand overschrijven: dit is het dossier dat naar
    # de accountant gaat en er zitten tabbladen in die ik niet zelf heb gemaakt.
    reserve = DOSSIER.with_name(DOSSIER.stem + " (kopie voor bijwerken).xlsx")
    shutil.copy2(DOSSIER, reserve)

    wb = openpyxl.load_workbook(DOSSIER)
    bestond = list(wb.sheetnames)
    for naam, bestand, uitleg in BLADEN:
        pad = MAP / bestand
        if not pad.exists():
            print(f"  overgeslagen (ontbreekt): {bestand}")
            continue
        n = zet_blad(wb, naam, pad, uitleg)
        merk = "vervangen" if naam in bestond else "toegevoegd"
        print(f"  {merk}: {naam} ({n} regels)")

    wb.save(DOSSIER)
    print(f"\ntabbladen nu: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print("  -", s)
    print(f"\nterugvalkopie: {reserve.name}")
    print(f"bijgewerkt op {date.today().strftime('%d-%m-%Y')}")


if __name__ == "__main__":
    main()
